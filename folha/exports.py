"""
Módulo de exportação de folha de pagamento
Suporta exportação para PDF e Excel
"""
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class FolhaPagamentoExporter:
    """Classe para exportação de folha de pagamento"""
    
    def __init__(self, folha):
        self.folha = folha
        self.resumos = folha.resumos.select_related('funcionario').order_by('funcionario__nome_completo')
    
    def export_pdf(self):
        """Exporta a folha para PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # Container para os elementos
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Título
        title = Paragraph(f"Folha de Pagamento - {self.folha.periodo_referencia}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))
        
        # Dados da tabela
        data = [
            ['Funcionário', 'Função', 'Proventos', 'Descontos', 'Líquido']
        ]
        
        for resumo in self.resumos:
            data.append([
                resumo.funcionario.nome_completo,
                resumo.funcionario.funcao.nome,
                f"R$ {resumo.total_proventos:,.2f}",
                f"R$ {resumo.total_descontos:,.2f}",
                f"R$ {resumo.valor_liquido:,.2f}",
            ])
        
        # Totais
        total_proventos = sum(r.total_proventos for r in self.resumos)
        total_descontos = sum(r.total_descontos for r in self.resumos)
        total_liquido = sum(r.valor_liquido for r in self.resumos)
        
        data.append([
            'TOTAL',
            '',
            f"R$ {total_proventos:,.2f}",
            f"R$ {total_descontos:,.2f}",
            f"R$ {total_liquido:,.2f}",
        ])
        
        # Cria a tabela
        table = Table(data, colWidths=[6*cm, 4*cm, 3*cm, 3*cm, 3*cm])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('GRID', (0, 0), (-1, -2), 1, colors.grey),
            
            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer = Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - Sistema de Folha de Pagamento Sonet 4.5",
            footer_style
        )
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    def export_excel(self):
        """Exporta a folha para Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Folha {self.folha.mes:02d}-{self.folha.ano}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="e0e7ff", end_color="e0e7ff", fill_type="solid")
        
        currency_alignment = Alignment(horizontal="right")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"FOLHA DE PAGAMENTO - {self.folha.periodo_referencia}"
        title_cell.font = Font(bold=True, size=14, color="1e40af")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Funcionário', 'Função', 'Proventos', 'Descontos', 'Líquido']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        row = 4
        for resumo in self.resumos:
            ws.cell(row=row, column=1, value=resumo.funcionario.nome_completo).border = border
            ws.cell(row=row, column=2, value=resumo.funcionario.funcao.nome).border = border
            
            cell_prov = ws.cell(row=row, column=3, value=float(resumo.total_proventos))
            cell_prov.number_format = 'R$ #,##0.00'
            cell_prov.alignment = currency_alignment
            cell_prov.border = border
            
            cell_desc = ws.cell(row=row, column=4, value=float(resumo.total_descontos))
            cell_desc.number_format = 'R$ #,##0.00'
            cell_desc.alignment = currency_alignment
            cell_desc.border = border
            
            cell_liq = ws.cell(row=row, column=5, value=float(resumo.valor_liquido))
            cell_liq.number_format = 'R$ #,##0.00'
            cell_liq.alignment = currency_alignment
            cell_liq.border = border
            
            row += 1
        
        # Totais
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill
        ws.cell(row=total_row, column=1).border = border
        ws.cell(row=total_row, column=2).fill = total_fill
        ws.cell(row=total_row, column=2).border = border
        
        total_proventos = sum(r.total_proventos for r in self.resumos)
        total_descontos = sum(r.total_descontos for r in self.resumos)
        total_liquido = sum(r.valor_liquido for r in self.resumos)
        
        cell_total_prov = ws.cell(row=total_row, column=3, value=float(total_proventos))
        cell_total_prov.font = total_font
        cell_total_prov.fill = total_fill
        cell_total_prov.number_format = 'R$ #,##0.00'
        cell_total_prov.alignment = currency_alignment
        cell_total_prov.border = border
        
        cell_total_desc = ws.cell(row=total_row, column=4, value=float(total_descontos))
        cell_total_desc.font = total_font
        cell_total_desc.fill = total_fill
        cell_total_desc.number_format = 'R$ #,##0.00'
        cell_total_desc.alignment = currency_alignment
        cell_total_desc.border = border
        
        cell_total_liq = ws.cell(row=total_row, column=5, value=float(total_liquido))
        cell_total_liq.font = total_font
        cell_total_liq.fill = total_fill
        cell_total_liq.number_format = 'R$ #,##0.00'
        cell_total_liq.alignment = currency_alignment
        cell_total_liq.border = border
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Salva em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer


def export_folha_pdf(folha):
    """Helper function para exportar folha em PDF"""
    exporter = FolhaPagamentoExporter(folha)
    buffer = exporter.export_pdf()
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f'folha_pagamento_{folha.ano}_{folha.mes:02d}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_folha_excel(folha):
    """Helper function para exportar folha em Excel"""
    exporter = FolhaPagamentoExporter(folha)
    buffer = exporter.export_excel()
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'folha_pagamento_{folha.ano}_{folha.mes:02d}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


class HoleriteExporter:
    """Classe para exportação de holerite individual"""
    
    def __init__(self, folha, funcionario):
        self.folha = folha
        self.funcionario = funcionario
        self.resumo = folha.resumos.filter(funcionario=funcionario).first()
        self.itens = folha.itens.filter(funcionario=funcionario).select_related('provento_desconto')
    
    def export_pdf(self):
        """Exporta o holerite para PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo customizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para subtítulo
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#4b5563'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Estilo para labels
        label_style = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=2
        )
        
        # Estilo para valores
        value_style = ParagraphStyle(
            'Value',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        # Cabeçalho
        title = Paragraph("CONTRACHEQUE / HOLERITE", title_style)
        elements.append(title)
        
        subtitle = Paragraph(
            f"Competência: {self.folha.periodo_referencia}",
            subtitle_style
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.5*cm))
        
        # Dados do Funcionário
        dados_funcionario = [
            ['DADOS DO FUNCIONÁRIO'],
        ]
        
        info_table_data = [
            [Paragraph('Nome:', label_style), Paragraph(self.funcionario.nome_completo, value_style)],
            [Paragraph('CPF:', label_style), Paragraph(self.funcionario.cpf, value_style)],
            [Paragraph('Função:', label_style), Paragraph(self.funcionario.funcao.nome, value_style)],
            [Paragraph('Setor:', label_style), Paragraph(self.funcionario.setor.nome, value_style)],
            [Paragraph('Admissão:', label_style), Paragraph(self.funcionario.data_admissao.strftime('%d/%m/%Y'), value_style)],
        ]
        
        # Tabela de cabeçalho
        header_table = Table(dados_funcionario, colWidths=[17*cm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
        ]))
        elements.append(header_table)
        
        # Tabela de informações
        info_table = Table(info_table_data, colWidths=[3*cm, 14*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Proventos
        proventos_data = [
            ['PROVENTOS', 'VALOR']
        ]
        
        proventos = self.itens.filter(provento_desconto__tipo='P').order_by('provento_desconto__nome')
        for item in proventos:
            proventos_data.append([
                item.provento_desconto.nome,
                f"R$ {item.valor_lancado:,.2f}"
            ])
        
        if len(proventos_data) > 1:
            proventos_table = Table(proventos_data, colWidths=[12*cm, 5*cm])
            proventos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecfdf5')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(proventos_table)
            elements.append(Spacer(1, 0.3*cm))
        
        # Descontos
        descontos_data = [
            ['DESCONTOS', 'VALOR']
        ]
        
        descontos = self.itens.filter(provento_desconto__tipo='D').order_by('provento_desconto__nome')
        for item in descontos:
            descontos_data.append([
                item.provento_desconto.nome,
                f"R$ {item.valor_lancado:,.2f}"
            ])
        
        if len(descontos_data) > 1:
            descontos_table = Table(descontos_data, colWidths=[12*cm, 5*cm])
            descontos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fee2e2')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(descontos_table)
            elements.append(Spacer(1, 0.5*cm))
        
        # Totalizadores
        if self.resumo:
            total_data = [
                ['Total de Proventos:', f"R$ {self.resumo.total_proventos:,.2f}"],
                ['Total de Descontos:', f"R$ {self.resumo.total_descontos:,.2f}"],
                ['VALOR LÍQUIDO:', f"R$ {self.resumo.valor_liquido:,.2f}"],
            ]
        else:
            total_proventos = sum(item.valor_lancado for item in proventos)
            total_descontos = sum(item.valor_lancado for item in descontos)
            valor_liquido = total_proventos - total_descontos
            
            total_data = [
                ['Total de Proventos:', f"R$ {total_proventos:,.2f}"],
                ['Total de Descontos:', f"R$ {total_descontos:,.2f}"],
                ['VALOR LÍQUIDO:', f"R$ {valor_liquido:,.2f}"],
            ]
        
        total_table = Table(total_data, colWidths=[12*cm, 5*cm])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#e0e7ff')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.whitesmoke),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica'),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 10),
            ('FONTSIZE', (0, 2), (-1, 2), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(total_table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer = Paragraph(
            f"Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} - Sistema de Folha de Pagamento Sonet 4.5",
            footer_style
        )
        elements.append(footer)
        
        # Aviso legal
        elements.append(Spacer(1, 0.3*cm))
        aviso_style = ParagraphStyle(
            'Aviso',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        aviso = Paragraph(
            "Este documento é apenas informativo e não substitui o contracheque oficial.",
            aviso_style
        )
        elements.append(aviso)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer


def export_holerite_pdf(folha, funcionario):
    """Helper function para exportar holerite individual em PDF"""
    exporter = HoleriteExporter(folha, funcionario)
    buffer = exporter.export_pdf()
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # Nome do arquivo sanitizado
    nome_limpo = ''.join(c for c in funcionario.nome_completo if c.isalnum() or c in (' ', '-', '_')).strip()
    nome_limpo = nome_limpo.replace(' ', '_')
    
    filename = f'holerite_{nome_limpo}_{folha.ano}_{folha.mes:02d}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


class RelatorioPagamentoExporter:
    """Classe para geração de relatórios de pagamento para conferência em papel e envio ao financeiro"""
    
    def __init__(self, folha, evento=None, filtro_pagamento='todos'):
        self.folha = folha
        self.evento = evento
        self.filtro_pagamento = filtro_pagamento

    def export_pdf(self):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.2*cm,
            leftMargin=1.2*cm,
            topMargin=1.2*cm,
            bottomMargin=1.2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilos Customizados
        title_style = ParagraphStyle(
            'RelatorioTitle',
            parent=styles['Heading1'],
            fontSize=15,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1e3a8a'),
            alignment=TA_CENTER,
            spaceAfter=3
        )
        subtitle_style = ParagraphStyle(
            'RelatorioSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#374151'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        body_style = ParagraphStyle(
            'RelatorioBody',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName='Helvetica',
            leading=10.5
        )
        body_bold = ParagraphStyle(
            'RelatorioBodyBold',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName='Helvetica-Bold',
            leading=10.5
        )
        center_style = ParagraphStyle(
            'RelatorioCenter',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName='Helvetica',
            alignment=TA_CENTER
        )
        right_style = ParagraphStyle(
            'RelatorioRight',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName='Helvetica',
            alignment=TA_RIGHT
        )
        right_bold = ParagraphStyle(
            'RelatorioRightBold',
            parent=styles['Normal'],
            fontSize=8.5,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT
        )
        
        # Título
        if self.evento:
            titulo_texto = f"RELATÓRIO DE PAGAMENTO - {self.evento.descricao.upper()}"
            subtitulo_texto = f"Competência: {self.folha.periodo_referencia} | Data Prevista: {self.evento.data_evento.strftime('%d/%m/%Y')}"
        else:
            titulo_texto = f"RELATÓRIO DE PAGAMENTO - CONSOLIDADO MENSAL"
            subtitulo_texto = f"Competência: {self.folha.periodo_referencia} | Folha de Pagamento Sonet 4.5"
            
        elements.append(Paragraph(titulo_texto, title_style))
        elements.append(Paragraph(subtitulo_texto, subtitle_style))
        
        # Preparação dos Dados conforme o tipo
        from decimal import Decimal
        linhas_dados = []
        total_a_pagar = Decimal('0.00')
        total_pago = Decimal('0.00')
        total_pendente = Decimal('0.00')
        qtd_total = 0
        qtd_pagos = 0
        
        if self.evento and self.evento.tipo_evento == 'AD':
            # Evento de Adiantamento Quinzenal
            from funcionarios.models import Adiantamento
            ads = Adiantamento.objects.filter(
                data_adiantamento__month=self.folha.mes,
                data_adiantamento__year=self.folha.ano
            ).select_related('funcionario', 'funcionario__funcao', 'funcionario__setor').order_by('funcionario__nome_completo')
            
            headers = ['Nº', 'Colaborador / Cargo / Setor', 'Chave PIX', 'Valor Adiantamento', 'Situação', 'Visto / Assinatura']
            col_widths = [1.2*cm, 7.5*cm, 5.0*cm, 3.5*cm, 3.5*cm, 6.0*cm]
            
            idx = 1
            for ad in ads:
                qtd_total += 1
                if ad.pago:
                    qtd_pagos += 1
                    total_pago += ad.valor
                else:
                    total_pendente += ad.valor
                total_a_pagar += ad.valor
                
                # Filtro
                if self.filtro_pagamento == 'pagos' and not ad.pago:
                    continue
                if self.filtro_pagamento == 'pendentes' and ad.pago:
                    continue
                
                colab_html = f"<b>{ad.funcionario.nome_completo}</b><br/><font size=7 color='#4b5563'>{ad.funcionario.funcao.nome if ad.funcionario.funcao else ''} • {ad.funcionario.setor.nome if ad.funcionario.setor else ''}</font>"
                pix_html = f"<font face='Courier' size=8>{ad.funcionario.chave_pix or 'Não cadastrado'}</font>"
                
                if ad.pago:
                    data_str = ad.data_pagamento.strftime('%d/%m/%Y') if ad.data_pagamento else ''
                    status_html = f"<font color='#15803d'><b>[PAGO]</b> {data_str}</font>"
                else:
                    status_html = "<font color='#b45309'><b>[PENDENTE]</b></font>"
                
                linhas_dados.append([
                    Paragraph(str(idx), center_style),
                    Paragraph(colab_html, body_style),
                    Paragraph(pix_html, body_style),
                    Paragraph(f"R$ {ad.valor:,.2f}", right_bold),
                    Paragraph(status_html, center_style),
                    Paragraph("<font color='#9ca3af'>___________________________</font>", center_style)
                ])
                idx += 1
                
            linha_total = [
                Paragraph("TOTAL", center_style),
                Paragraph(f"<b>{len(linhas_dados)} colaborador(es) listado(s)</b>", body_bold),
                "",
                Paragraph(f"<b>R$ {total_a_pagar:,.2f}</b>", right_bold),
                "",
                ""
            ]
        elif self.evento and self.evento.tipo_evento == 'PF':
            # Pagamento Final do Mês
            resumos = self.folha.resumos.select_related('funcionario', 'funcionario__funcao', 'funcionario__setor').order_by('funcionario__nome_completo')
            headers = ['Nº', 'Colaborador / Cargo', 'Chave PIX', 'Salário Bruto', 'Desc. Adiant.', 'Líquido a Pagar', 'Situação', 'Visto / Assinatura']
            col_widths = [1.0*cm, 5.8*cm, 4.2*cm, 3.0*cm, 2.8*cm, 3.2*cm, 2.7*cm, 4.0*cm]
            
            idx = 1
            for r in resumos:
                qtd_total += 1
                if r.pago:
                    qtd_pagos += 1
                    total_pago += r.valor_liquido
                else:
                    total_pendente += r.valor_liquido
                total_a_pagar += r.valor_liquido
                
                if self.filtro_pagamento == 'pagos' and not r.pago:
                    continue
                if self.filtro_pagamento == 'pendentes' and r.pago:
                    continue
                
                colab_html = f"<b>{r.funcionario.nome_completo}</b><br/><font size=7 color='#4b5563'>{r.funcionario.funcao.nome if r.funcionario.funcao else ''}</font>"
                pix_html = f"<font face='Courier' size=7.5>{r.funcionario.chave_pix or 'Não informado'}</font>"
                
                if r.pago:
                    data_str = r.data_pagamento.strftime('%d/%m') if r.data_pagamento else ''
                    status_html = f"<font color='#15803d'><b>[PAGO]</b> {data_str}</font>"
                else:
                    status_html = "<font color='#b45309'><b>[PEND.]</b></font>"
                
                linhas_dados.append([
                    Paragraph(str(idx), center_style),
                    Paragraph(colab_html, body_style),
                    Paragraph(pix_html, body_style),
                    Paragraph(f"R$ {r.total_proventos:,.2f}", right_style),
                    Paragraph(f"R$ {r.total_descontos:,.2f}", right_style),
                    Paragraph(f"R$ {r.valor_liquido:,.2f}", right_bold),
                    Paragraph(status_html, center_style),
                    Paragraph("<font color='#9ca3af'>___________________</font>", center_style)
                ])
                idx += 1
                
            linha_total = [
                Paragraph("TOTAL", center_style),
                Paragraph(f"<b>{len(linhas_dados)} listado(s)</b>", body_bold),
                "",
                Paragraph(f"R$ {sum(r.total_proventos for r in resumos):,.2f}", right_bold),
                Paragraph(f"R$ {sum(r.total_descontos for r in resumos):,.2f}", right_bold),
                Paragraph(f"<b>R$ {total_a_pagar:,.2f}</b>", right_bold),
                "",
                ""
            ]
        else:
            # Visão Consolidada Geral
            resumos = self.folha.resumos.select_related('funcionario', 'funcionario__funcao', 'funcionario__setor').order_by('funcionario__nome_completo')
            headers = ['Nº', 'Colaborador', 'Cargo / Setor', 'Chave PIX', 'Total Proventos', 'Total Descontos', 'Total Líquido', 'Status']
            col_widths = [1.2*cm, 6.0*cm, 4.5*cm, 4.5*cm, 3.2*cm, 3.2*cm, 3.5*cm, 2.6*cm]
            
            idx = 1
            for r in resumos:
                qtd_total += 1
                if r.pago:
                    qtd_pagos += 1
                    total_pago += r.valor_liquido
                else:
                    total_pendente += r.valor_liquido
                total_a_pagar += r.valor_liquido
                
                if self.filtro_pagamento == 'pagos' and not r.pago:
                    continue
                if self.filtro_pagamento == 'pendentes' and r.pago:
                    continue
                
                colab_html = f"<b>{r.funcionario.nome_completo}</b>"
                funcao_html = f"{r.funcionario.funcao.nome if r.funcionario.funcao else ''}<br/><font size=7 color='#4b5563'>{r.funcionario.setor.nome if r.funcionario.setor else ''}</font>"
                pix_html = f"<font face='Courier' size=8>{r.funcionario.chave_pix or 'Não cadastrado'}</font>"
                
                if r.pago:
                    status_html = "<font color='#15803d'><b>PAGO</b></font>"
                else:
                    status_html = "<font color='#b45309'><b>PENDENTE</b></font>"
                
                linhas_dados.append([
                    Paragraph(str(idx), center_style),
                    Paragraph(colab_html, body_style),
                    Paragraph(funcao_html, body_style),
                    Paragraph(pix_html, body_style),
                    Paragraph(f"R$ {r.total_proventos:,.2f}", right_style),
                    Paragraph(f"R$ {r.total_descontos:,.2f}", right_style),
                    Paragraph(f"R$ {r.valor_liquido:,.2f}", right_bold),
                    Paragraph(status_html, center_style)
                ])
                idx += 1
                
            linha_total = [
                Paragraph("TOTAL", center_style),
                Paragraph(f"<b>{len(linhas_dados)} colaborador(es)</b>", body_bold),
                "",
                "",
                Paragraph(f"R$ {sum(r.total_proventos for r in resumos):,.2f}", right_bold),
                Paragraph(f"R$ {sum(r.total_descontos for r in resumos):,.2f}", right_bold),
                Paragraph(f"<b>R$ {total_a_pagar:,.2f}</b>", right_bold),
                ""
            ]

        # Card de Resumo no Topo do Relatório
        resumo_card_data = [
            [
                Paragraph(f"<b>Total a Desembolsar:</b> R$ {total_a_pagar:,.2f}", body_style),
                Paragraph(f"<b>Já Pago:</b> <font color='#15803d'>R$ {total_pago:,.2f}</font> ({qtd_pagos})", body_style),
                Paragraph(f"<b>Pendente:</b> <font color='#b45309'>R$ {total_pendente:,.2f}</font> ({qtd_total - qtd_pagos})", body_style),
                Paragraph(f"<b>Filtro:</b> {self.filtro_pagamento.upper()}", body_style),
                Paragraph(f"<b>Emissão:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", right_style),
            ]
        ]
        resumo_card = Table(resumo_card_data, colWidths=[6.0*cm, 5.5*cm, 5.5*cm, 4.0*cm, 5.7*cm])
        resumo_card.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f3f4f6')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(resumo_card)
        elements.append(Spacer(1, 0.4*cm))

        # Montagem da Tabela Principal
        tabela_dados = [[Paragraph(f"<b>{h}</b>", center_style) for h in headers]]
        for row in linhas_dados:
            tabela_dados.append(row)
        tabela_dados.append(linha_total)

        tabela_principal = Table(tabela_dados, colWidths=col_widths, repeatRows=1)
        
        # Estilização da Tabela
        tabela_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#9ca3af')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#1e40af')),
            ('TOPPADDING', (0, -1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
        ]
        
        # Zebra striping
        for i in range(1, len(tabela_dados) - 1):
            if i % 2 == 0:
                tabela_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f9fafb')))
            tabela_style.append(('TOPPADDING', (0, i), (-1, i), 5))
            tabela_style.append(('BOTTOMPADDING', (0, i), (-1, i), 5))

        tabela_principal.setStyle(TableStyle(tabela_style))
        elements.append(tabela_principal)

        # Bloco de Assinaturas para Conferência
        elements.append(Spacer(1, 1.0*cm))
        assinaturas_data = [
            [
                Paragraph("<font color='#6b7280'>____________________________________________<br/><b>Elaborado por (RH / DP)</b></font>", center_style),
                Paragraph("<font color='#6b7280'>____________________________________________<br/><b>Conferido / Pago por (Financeiro)</b></font>", center_style),
                Paragraph("<font color='#6b7280'>____________________________________________<br/><b>Autorizado por (Diretoria)</b></font>", center_style),
            ]
        ]
        assinaturas_table = Table(assinaturas_data, colWidths=[9.0*cm, 9.0*cm, 9.0*cm])
        assinaturas_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(assinaturas_table)

        # Rodapé
        elements.append(Spacer(1, 0.4*cm))
        footer_style = ParagraphStyle(
            'RelatorioFooter',
            parent=styles['Normal'],
            fontSize=7.5,
            textColor=colors.HexColor('#9ca3af'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} • Sistema de Folha de Pagamento Sonet 4.5 • Documento para uso interno e conferência bancária",
            footer_style
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer


def export_relatorio_pagamento_pdf(folha, evento=None, filtro_pagamento='todos'):
    """Helper function para exportar o relatório de pagamento para o financeiro em PDF"""
    exporter = RelatorioPagamentoExporter(folha, evento=evento, filtro_pagamento=filtro_pagamento)
    buffer = exporter.export_pdf()
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    if evento:
        slug_evento = ''.join(c for c in evento.descricao if c.isalnum() or c in (' ', '-', '_')).strip().replace(' ', '_')
        filename = f"relatorio_{slug_evento}_{folha.ano}_{folha.mes:02d}_{filtro_pagamento}.pdf"
    else:
        filename = f"relatorio_folha_consolidada_{folha.ano}_{folha.mes:02d}_{filtro_pagamento}.pdf"
        
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response
